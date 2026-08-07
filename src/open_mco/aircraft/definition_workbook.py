"""Importer for the minimal Aircraft One spreadsheet contract."""

from __future__ import annotations

import hashlib
import re
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from .loader import AircraftWorkbookError
from .specification import (
    AircraftDefinition,
    AircraftField,
    NearFieldSample,
    PerformancePoint,
    PhasePoint,
    PhaseTiming,
)

PARAMETER_HEADERS = [
    "Parameter",
    "Value",
    "Unit",
    "Required",
    "Evidence Class",
    "Source",
    "Source URL",
    "Page/Figure",
    "Notes",
]
PARAMETER_SHEETS = {
    "General": "General",
    "Operating_Limits": "Operating Limits",
    "Mission_Config": "Mission Config",
    "Sonic_Boom": "Sonic Boom",
}
REQUIRED_SHEETS = {
    "README",
    *PARAMETER_SHEETS,
    "Performance_Map",
    "Phase_Profile",
    "Phase_Timing",
    "Nearfield_Signatures",
}
LEGACY_PARAMETER_HEADERS = [
    "Parameter",
    "Value",
    "Unit",
    "Required",
    "Source",
    "Page/Figure",
    "Notes",
]
LEGACY_PARAMETER_SHEETS = {
    "General": "General",
    "Operating_Limits": "Operating Limits",
    "Mission_Config": "Mission Config",
    "Sonic_Boom_Optional": "Sonic Boom",
}
LEGACY_REQUIRED_SHEETS = {
    "General",
    "Operating_Limits",
    "Performance_Map",
    "Mission_Config",
    "Phase_Profile",
    "Phase_Timing",
    "Sonic_Boom_Optional",
}
PERFORMANCE_HEADERS = [
    "Weight (lb)",
    "Altitude (ft)",
    "Mach",
    "Throttle (%)",
    "Drag Required (lbf)",
    "Available Thrust (lbf)",
    "Fuel Flow (lb/hr)",
    "Sustainable (Y/N)",
    "Evidence Class",
    "Source",
    "Source URL",
    "Page/Figure",
    "Notes",
]
PHASE_HEADERS = [
    "Sequence",
    "Phase",
    "Altitude (ft)",
    "Mach",
    "Evidence Class",
    "Source",
    "Source URL",
    "Page/Figure",
    "Notes",
]
TIMING_HEADERS = [
    "Phase",
    "Duration (min)",
    "Basis",
    "Source",
    "Source URL",
    "Page/Figure",
    "Notes",
]
NEARFIELD_HEADERS = [
    "Signature ID",
    "Axial Position (ft)",
    "Delta Pressure (psf)",
    "Reference Distance (ft)",
    "Azimuth (deg)",
    "Mach",
    "Altitude (ft)",
    "Weight (lb)",
    "Angle of Attack (deg)",
    "Evidence Class",
    "Source",
    "Source URL",
    "Page/Figure",
    "Notes",
]
LEGACY_PHASE_HEADERS = [
    "Sequence",
    "Phase",
    "Altitude (ft)",
    "Mach",
    "Source",
    "Page/Figure",
    "Notes",
]
LEGACY_TIMING_HEADERS = [
    "Phase",
    "Duration (min)",
    "Basis",
    "Source",
    "Page/Figure",
    "Notes",
]


def _bytes(source: str | Path | BinaryIO | bytes) -> bytes:
    if isinstance(source, bytes):
        return source
    if isinstance(source, (str, Path)):
        return Path(source).read_bytes()
    payload = source.read()
    if not isinstance(payload, bytes):
        raise AircraftWorkbookError("uploaded aircraft workbook did not provide binary data")
    return payload


def _text(value: Any) -> str | None:
    if value is None or str(value).strip() == "":
        return None
    return str(value).strip()


def _headers(sheet: Any, expected: list[str]) -> None:
    actual = [cell.value for cell in sheet[1]][: len(expected)]
    if actual != expected:
        raise AircraftWorkbookError(
            f"sheet {sheet.title!r} must use columns: {' | '.join(expected)}"
        )


def _records(sheet: Any, headers: list[str]) -> list[dict[str, Any]]:
    _headers(sheet, headers)
    return [
        dict(zip(headers, row[: len(headers)], strict=True))
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if any(value not in (None, "") for value in row[: len(headers)])
    ]


def _numeric(row: dict[str, Any], key: str, sheet: str, index: int) -> float:
    try:
        return float(row[key])
    except (TypeError, ValueError) as exc:
        raise AircraftWorkbookError(f"{sheet} row {index} has invalid {key}") from exc


def _traceability(row: dict[str, Any], sheet: str, index: int) -> None:
    evidence = (_text(row.get("Evidence Class")) or "UNAVAILABLE").upper()
    if evidence == "PUBLISHED" and (
        not _text(row.get("Source"))
        or not _text(row.get("Source URL"))
        or not _text(row.get("Page/Figure"))
    ):
        raise AircraftWorkbookError(
            f"{sheet} row {index} is PUBLISHED but lacks source, URL, or page/figure"
        )


def _split_legacy_source(value: Any) -> tuple[str | None, str | None]:
    lines = [line.strip() for line in str(value or "").splitlines() if line.strip()]
    url = next((line for line in lines if line.startswith(("http://", "https://"))), None)
    names = [line for line in lines if line != url]
    return (" · ".join(names) or None, url)


def _legacy_evidence(value: Any, source: Any, page: Any) -> str:
    if value in (None, ""):
        return "UNAVAILABLE"
    source_name, source_url = _split_legacy_source(source)
    return (
        "PUBLISHED"
        if source_name and source_url and _text(page)
        else "UNVALIDATED_ASSUMPTION"
    )


def _legacy_parameter_fields(workbook: Any) -> list[AircraftField]:
    fields: list[AircraftField] = []
    for sheet_name, section in LEGACY_PARAMETER_SHEETS.items():
        for row in _records(workbook[sheet_name], LEGACY_PARAMETER_HEADERS):
            parameter = _text(row["Parameter"])
            if not parameter:
                continue
            source_name, source_url = _split_legacy_source(row["Source"])
            fields.append(
                AircraftField(
                    section=section,
                    parameter=parameter,
                    value=_text(row["Value"]),
                    unit=_text(row["Unit"]) or "-",
                    required=(_text(row["Required"]) or "No").lower() == "yes",
                    evidence_class=_legacy_evidence(
                        row["Value"], row["Source"], row["Page/Figure"]
                    ),
                    source_name=source_name,
                    source_url=source_url,
                    page_figure=_text(row["Page/Figure"]),
                    notes=_text(row["Notes"]),
                )
            )
    existing = {field.parameter for field in fields}
    for parameter, notes in (
        ("Propagation Engine", "Nonlinear propagation implementation and configuration."),
        ("Propagation Engine Version", None),
        ("Primary Ray Handling", None),
        ("Secondary Ray Handling", None),
        ("Ground Metrics", "Ground waveform, peak overpressure, and selected loudness metrics."),
        ("PCBoom Validation Case", "PCBoom version, configuration, case, and comparison outputs."),
    ):
        if parameter not in existing:
            fields.append(
                AircraftField(
                    section="Sonic Boom",
                    parameter=parameter,
                    value=None,
                    unit="-",
                    required=True,
                    evidence_class="UNAVAILABLE",
                    notes=notes,
                )
            )
    return fields


def _legacy_phase_profile(workbook: Any) -> tuple[PhasePoint, ...]:
    points: list[PhasePoint] = []
    for index, row in enumerate(
        _records(workbook["Phase_Profile"], LEGACY_PHASE_HEADERS), start=2
    ):
        if any(row[key] in (None, "") for key in LEGACY_PHASE_HEADERS[:6]):
            raise AircraftWorkbookError(
                f"Phase_Profile row {index} is partial; complete it or remove it"
            )
        source_name, source_url = _split_legacy_source(row["Source"])
        points.append(
            PhasePoint(
                sequence=int(_numeric(row, "Sequence", "Phase_Profile", index)),
                phase=_text(row["Phase"]) or "",
                altitude_ft=_numeric(row, "Altitude (ft)", "Phase_Profile", index),
                mach=_numeric(row, "Mach", "Phase_Profile", index),
                evidence_class=_legacy_evidence(
                    row["Mach"], row["Source"], row["Page/Figure"]
                ),
                source_name=source_name or "UNSPECIFIED",
                source_url=source_url or "UNSPECIFIED",
                page_figure=_text(row["Page/Figure"]) or "UNSPECIFIED",
                notes=_text(row["Notes"]),
            )
        )
    if [point.sequence for point in points] != list(range(1, len(points) + 1)):
        raise AircraftWorkbookError("Phase_Profile Sequence must be contiguous from 1")
    return tuple(points)


def _legacy_phase_timing(workbook: Any) -> tuple[PhaseTiming, ...]:
    timings: list[PhaseTiming] = []
    for index, row in enumerate(
        _records(workbook["Phase_Timing"], LEGACY_TIMING_HEADERS), start=2
    ):
        phase = _text(row["Phase"])
        if not phase:
            raise AircraftWorkbookError(f"Phase_Timing row {index} has no phase")
        source_name, source_url = _split_legacy_source(row["Source"])
        timings.append(
            PhaseTiming(
                phase=phase,  # type: ignore[arg-type]
                duration_min=(
                    None
                    if row["Duration (min)"] in (None, "")
                    else _numeric(row, "Duration (min)", "Phase_Timing", index)
                ),
                basis=(_text(row["Basis"]) or "MISSING").upper(),  # type: ignore[arg-type]
                source_name=source_name,
                source_url=source_url,
                page_figure=_text(row["Page/Figure"]),
                notes=_text(row["Notes"]),
            )
        )
    return tuple(timings)


def _legacy_has_performance_rows(workbook: Any) -> bool:
    sheet = workbook["Performance_Map"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = row[:9]
        if not any(value not in (None, "") for value in values):
            continue
        if all(isinstance(value, str) for value in values[:4]):
            continue
        return True
    return False


def _load_legacy_definition(workbook: Any, payload: bytes) -> AircraftDefinition:
    missing = LEGACY_REQUIRED_SHEETS - set(workbook.sheetnames)
    if missing:
        raise AircraftWorkbookError(f"missing required sheets: {', '.join(sorted(missing))}")
    if _legacy_has_performance_rows(workbook):
        raise AircraftWorkbookError(
            "legacy Performance_Map rows lack throttle and drag; move them into the current "
            "template so MachLane does not invent those values"
        )
    fields = _legacy_parameter_fields(workbook)
    aircraft_name = next(
        (field.value for field in fields if field.parameter == "Aircraft Name"), None
    )
    return AircraftDefinition(
        aircraft_id="aircraft_one",
        display_name=aircraft_name or "Uploaded Aircraft",
        revision=1,
        updated_at=datetime.now(UTC),
        fields=tuple(fields),
        phase_profile=_legacy_phase_profile(workbook),
        phase_timing=_legacy_phase_timing(workbook),
        performance_map=(),
        nearfield_samples=(),
        workbook_checksum=hashlib.sha256(payload).hexdigest(),
    )


def _parameter_fields(workbook: Any) -> tuple[list[AircraftField], dict[str, Any]]:
    fields: list[AircraftField] = []
    metadata: dict[str, Any] = {}
    for sheet_name, section in PARAMETER_SHEETS.items():
        sheet = workbook[sheet_name]
        for index, row in enumerate(_records(sheet, PARAMETER_HEADERS), start=2):
            parameter = _text(row["Parameter"])
            if not parameter:
                continue
            _traceability(row, sheet_name, index)
            if sheet_name == "General" and parameter in {"Aircraft ID", "Display Name"}:
                metadata[parameter] = row["Value"]
                continue
            required = (_text(row["Required"]) or "No").lower() == "yes"
            fields.append(
                AircraftField(
                    section=section,
                    parameter=parameter,
                    value=_text(row["Value"]),
                    unit=_text(row["Unit"]) or "-",
                    required=required,
                    evidence_class=(_text(row["Evidence Class"]) or "UNAVAILABLE").upper(),
                    source_name=_text(row["Source"]),
                    source_url=_text(row["Source URL"]),
                    page_figure=_text(row["Page/Figure"]),
                    notes=_text(row["Notes"]),
                )
            )
    return fields, metadata


def _performance_points(workbook: Any) -> tuple[PerformancePoint, ...]:
    points: list[PerformancePoint] = []
    for index, row in enumerate(
        _records(workbook["Performance_Map"], PERFORMANCE_HEADERS), start=2
    ):
        if any(row[key] in (None, "") for key in PERFORMANCE_HEADERS[:12]):
            raise AircraftWorkbookError(
                f"Performance_Map row {index} is partial; complete it or remove it"
            )
        _traceability(row, "Performance_Map", index)
        flag = (_text(row["Sustainable (Y/N)"]) or "").upper()
        if flag not in {"Y", "N"}:
            raise AircraftWorkbookError(
                f"Performance_Map row {index} Sustainable (Y/N) must be Y or N"
            )
        points.append(
            PerformancePoint(
                weight_lb=_numeric(row, "Weight (lb)", "Performance_Map", index),
                altitude_ft=_numeric(row, "Altitude (ft)", "Performance_Map", index),
                mach=_numeric(row, "Mach", "Performance_Map", index),
                throttle_percent=_numeric(row, "Throttle (%)", "Performance_Map", index),
                drag_required_lbf=_numeric(
                    row, "Drag Required (lbf)", "Performance_Map", index
                ),
                available_thrust_lbf=_numeric(
                    row, "Available Thrust (lbf)", "Performance_Map", index
                ),
                fuel_flow_lb_hr=_numeric(row, "Fuel Flow (lb/hr)", "Performance_Map", index),
                sustainable=flag == "Y",
                evidence_class=(_text(row["Evidence Class"]) or "UNAVAILABLE").upper(),
                source_name=_text(row["Source"]) or "UNSPECIFIED",
                source_url=_text(row["Source URL"]) or "UNSPECIFIED",
                page_figure=_text(row["Page/Figure"]) or "UNSPECIFIED",
                notes=_text(row["Notes"]),
            )
        )
    return tuple(points)


def _phase_profile(workbook: Any) -> tuple[PhasePoint, ...]:
    points: list[PhasePoint] = []
    for index, row in enumerate(_records(workbook["Phase_Profile"], PHASE_HEADERS), start=2):
        if any(row[key] in (None, "") for key in PHASE_HEADERS[:8]):
            raise AircraftWorkbookError(
                f"Phase_Profile row {index} is partial; complete it or remove it"
            )
        _traceability(row, "Phase_Profile", index)
        points.append(
            PhasePoint(
                sequence=int(_numeric(row, "Sequence", "Phase_Profile", index)),
                phase=_text(row["Phase"]) or "",
                altitude_ft=_numeric(row, "Altitude (ft)", "Phase_Profile", index),
                mach=_numeric(row, "Mach", "Phase_Profile", index),
                evidence_class=(_text(row["Evidence Class"]) or "UNAVAILABLE").upper(),
                source_name=_text(row["Source"]) or "UNSPECIFIED",
                source_url=_text(row["Source URL"]) or "UNSPECIFIED",
                page_figure=_text(row["Page/Figure"]) or "UNSPECIFIED",
                notes=_text(row["Notes"]),
            )
        )
    if [point.sequence for point in points] != list(range(1, len(points) + 1)):
        raise AircraftWorkbookError("Phase_Profile Sequence must be contiguous from 1")
    return tuple(points)


def _phase_timing(workbook: Any) -> tuple[PhaseTiming, ...]:
    timings: list[PhaseTiming] = []
    for index, row in enumerate(_records(workbook["Phase_Timing"], TIMING_HEADERS), start=2):
        phase = _text(row["Phase"])
        basis = (_text(row["Basis"]) or "MISSING").upper()
        if not phase:
            raise AircraftWorkbookError(f"Phase_Timing row {index} has no phase")
        duration = (
            None
            if row["Duration (min)"] in (None, "")
            else _numeric(row, "Duration (min)", "Phase_Timing", index)
        )
        timings.append(
            PhaseTiming(
                phase=phase,  # type: ignore[arg-type]
                duration_min=duration,
                basis=basis,  # type: ignore[arg-type]
                source_name=_text(row["Source"]),
                source_url=_text(row["Source URL"]),
                page_figure=_text(row["Page/Figure"]),
                notes=_text(row["Notes"]),
            )
        )
    return tuple(timings)


def _nearfield_samples(workbook: Any) -> tuple[NearFieldSample, ...]:
    samples: list[NearFieldSample] = []
    for index, row in enumerate(
        _records(workbook["Nearfield_Signatures"], NEARFIELD_HEADERS), start=2
    ):
        if any(row[key] in (None, "") for key in NEARFIELD_HEADERS[:13]):
            raise AircraftWorkbookError(
                f"Nearfield_Signatures row {index} is partial; complete it or remove it"
            )
        _traceability(row, "Nearfield_Signatures", index)
        samples.append(
            NearFieldSample(
                signature_id=_text(row["Signature ID"]) or "",
                axial_position_ft=_numeric(
                    row, "Axial Position (ft)", "Nearfield_Signatures", index
                ),
                delta_pressure_psf=_numeric(
                    row, "Delta Pressure (psf)", "Nearfield_Signatures", index
                ),
                reference_distance_ft=_numeric(
                    row, "Reference Distance (ft)", "Nearfield_Signatures", index
                ),
                azimuth_deg=_numeric(row, "Azimuth (deg)", "Nearfield_Signatures", index),
                mach=_numeric(row, "Mach", "Nearfield_Signatures", index),
                altitude_ft=_numeric(row, "Altitude (ft)", "Nearfield_Signatures", index),
                weight_lb=_numeric(row, "Weight (lb)", "Nearfield_Signatures", index),
                angle_of_attack_deg=_numeric(
                    row, "Angle of Attack (deg)", "Nearfield_Signatures", index
                ),
                evidence_class=(_text(row["Evidence Class"]) or "UNAVAILABLE").upper(),
                source_name=_text(row["Source"]) or "UNSPECIFIED",
                source_url=_text(row["Source URL"]) or "UNSPECIFIED",
                page_figure=_text(row["Page/Figure"]) or "UNSPECIFIED",
                notes=_text(row["Notes"]),
            )
        )
    return tuple(samples)


def load_aircraft_definition_workbook(
    source: str | Path | BinaryIO | bytes,
) -> AircraftDefinition:
    """Auto-detect a current or legacy aircraft workbook and normalize it."""

    payload = _bytes(source)
    try:
        workbook = load_workbook(BytesIO(payload), data_only=True, read_only=True)
    except Exception as exc:  # openpyxl raises several format-specific exception classes
        raise AircraftWorkbookError(f"aircraft workbook could not be opened: {exc}") from exc
    if "Sonic_Boom_Optional" in workbook.sheetnames:
        return _load_legacy_definition(workbook, payload)
    missing = REQUIRED_SHEETS - set(workbook.sheetnames)
    if missing:
        raise AircraftWorkbookError(f"missing required sheets: {', '.join(sorted(missing))}")
    fields, metadata = _parameter_fields(workbook)
    aircraft_id = (_text(metadata.get("Aircraft ID")) or "aircraft_one").lower()
    aircraft_id = re.sub(r"[^a-z0-9_-]+", "_", aircraft_id).strip("_")
    if aircraft_id != "aircraft_one":
        raise AircraftWorkbookError("Aircraft ID must be exactly 'aircraft_one'")
    display_name = _text(metadata.get("Display Name")) or "Aircraft One"
    if display_name != "Aircraft One":
        raise AircraftWorkbookError("Display Name must be exactly 'Aircraft One'")
    return AircraftDefinition(
        aircraft_id="aircraft_one",
        display_name="Aircraft One",
        revision=1,
        updated_at=datetime.now(UTC),
        fields=tuple(fields),
        phase_profile=_phase_profile(workbook),
        phase_timing=_phase_timing(workbook),
        performance_map=_performance_points(workbook),
        nearfield_samples=_nearfield_samples(workbook),
        workbook_checksum=hashlib.sha256(payload).hexdigest(),
    )


def export_aircraft_definition_workbook(definition: AircraftDefinition) -> bytes:
    """Serialize the normalized aircraft model to MachLane's current Excel contract."""

    workbook = Workbook()
    initial_sheet = workbook.active
    if initial_sheet is not None:
        workbook.remove(initial_sheet)
    readme = workbook.create_sheet("README")
    readme.append(["MachLane normalized aircraft workbook"])
    readme.append(["Aircraft", definition.value("Aircraft Name") or definition.display_name])
    readme.append(["Schema", definition.schema_version])
    readme.append(["Revision", definition.revision])
    readme.append(["Updated UTC", definition.updated_at.isoformat()])
    readme.append(
        [
            "Safety gate",
            "Surface boom remains unavailable until a reviewed near-field signature and validated propagation model are supplied.",
        ]
    )

    fields_by_section: dict[str, list[AircraftField]] = {
        section: [] for section in PARAMETER_SHEETS.values()
    }
    for field in definition.fields:
        if field.section in fields_by_section:
            fields_by_section[field.section].append(field)
    for sheet_name, section in PARAMETER_SHEETS.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(PARAMETER_HEADERS)
        if sheet_name == "General":
            sheet.append(
                [
                    "Aircraft ID",
                    "aircraft_one",
                    "-",
                    "Yes",
                    "CALCULATED",
                    "MachLane normalized workbook",
                    "https://github.com/ldc555/machlane",
                    definition.schema_version,
                    "Fixed importer identifier.",
                ]
            )
            sheet.append(
                [
                    "Display Name",
                    "Aircraft One",
                    "-",
                    "Yes",
                    "CALCULATED",
                    "MachLane normalized workbook",
                    "https://github.com/ldc555/machlane",
                    definition.schema_version,
                    "Fixed importer display name.",
                ]
            )
        for field in fields_by_section[section]:
            sheet.append(
                [
                    field.parameter,
                    field.value,
                    field.unit,
                    "Yes" if field.required else "No",
                    field.evidence_class,
                    field.source_name,
                    field.source_url,
                    field.page_figure,
                    field.notes,
                ]
            )

    performance = workbook.create_sheet("Performance_Map")
    performance.append(PERFORMANCE_HEADERS)
    for performance_point in definition.performance_map:
        performance.append(
            [
                performance_point.weight_lb,
                performance_point.altitude_ft,
                performance_point.mach,
                performance_point.throttle_percent,
                performance_point.drag_required_lbf,
                performance_point.available_thrust_lbf,
                performance_point.fuel_flow_lb_hr,
                "Y" if performance_point.sustainable else "N",
                performance_point.evidence_class,
                performance_point.source_name,
                performance_point.source_url,
                performance_point.page_figure,
                performance_point.notes,
            ]
        )

    phases = workbook.create_sheet("Phase_Profile")
    phases.append(PHASE_HEADERS)
    for phase_point in definition.phase_profile:
        phases.append(
            [
                phase_point.sequence,
                phase_point.phase,
                phase_point.altitude_ft,
                phase_point.mach,
                phase_point.evidence_class,
                phase_point.source_name,
                phase_point.source_url,
                phase_point.page_figure,
                phase_point.notes,
            ]
        )

    timing = workbook.create_sheet("Phase_Timing")
    timing.append(TIMING_HEADERS)
    for timing_point in definition.phase_timing:
        timing.append(
            [
                timing_point.phase,
                timing_point.duration_min,
                timing_point.basis,
                timing_point.source_name,
                timing_point.source_url,
                timing_point.page_figure,
                timing_point.notes,
            ]
        )

    nearfield = workbook.create_sheet("Nearfield_Signatures")
    nearfield.append(NEARFIELD_HEADERS)
    for sample in definition.nearfield_samples:
        nearfield.append(
            [
                sample.signature_id,
                sample.axial_position_ft,
                sample.delta_pressure_psf,
                sample.reference_distance_ft,
                sample.azimuth_deg,
                sample.mach,
                sample.altitude_ft,
                sample.weight_lb,
                sample.angle_of_attack_deg,
                sample.evidence_class,
                sample.source_name,
                sample.source_url,
                sample.page_figure,
                sample.notes,
            ]
        )

    header_fill = PatternFill("solid", fgColor="0F6173")
    for sheet in workbook.worksheets[1:]:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        sheet.auto_filter.ref = sheet.dimensions
    readme.column_dimensions["A"].width = 22
    readme.column_dimensions["B"].width = 100

    payload = BytesIO()
    workbook.save(payload)
    return payload.getvalue()
