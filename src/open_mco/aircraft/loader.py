"""Resilient loader for the versioned MachLane aircraft workbook contract."""

from __future__ import annotations

import hashlib
import warnings
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from open_mco.models import (
    AircraftModel,
    AircraftOperatingLimits,
    AircraftPerformancePoint,
    SourcedValue,
)
from open_mco.units import to_si

REQUIRED_SHEETS = {"General", "Operating_Limits", "Performance_Map", "Mission_Config"}
OPTIONAL_SHEETS = {"Aerodynamics_Optional", "Sonic_Boom_Optional"}


class AircraftWorkbookError(ValueError):
    """Raised when a workbook violates the published data contract."""


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _parameter_rows(sheet: Any) -> dict[str, dict[str, Any]]:
    headers = [cell.value for cell in sheet[1]]
    expected = ["Parameter", "Value", "Unit", "Required", "Source", "Page/Figure", "Notes"]
    if headers[:7] != expected:
        raise AircraftWorkbookError(
            f"sheet {sheet.title!r} must use columns: {' | '.join(expected)}"
        )
    output: dict[str, dict[str, Any]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        key = str(row[0]).strip()
        if key in output:
            raise AircraftWorkbookError(f"duplicate parameter {key!r} in {sheet.title}")
        output[key] = dict(zip(expected[1:], row[1:7], strict=True))
    return output


def _sourced(
    row: dict[str, Any], *, parameter: str, workbook: Path, checksum: str, retrieved_at: datetime
) -> SourcedValue | None:
    value = row["Value"]
    required = str(row["Required"] or "").strip().lower() == "yes"
    if value is None or value == "":
        if required:
            raise AircraftWorkbookError(f"required value {parameter!r} is blank")
        return None
    unit = str(row["Unit"] or "-").strip()
    try:
        value_si, si_unit = to_si(value, unit)
    except ValueError as exc:
        raise AircraftWorkbookError(f"{parameter}: {exc}") from exc
    source = str(row["Source"] or "").strip()
    page = str(row["Page/Figure"] or "").strip()
    if not source or not page:
        warnings.warn(
            f"{parameter!r} is populated without complete source and page/figure traceability",
            stacklevel=2,
        )
    return SourcedValue(
        original_value=value,
        original_unit=unit,
        value_si=value_si,
        si_unit=si_unit,
        source_name=source or "UNSPECIFIED",
        source_document=str(workbook),
        page_figure=page or None,
        retrieved_at=retrieved_at,
        checksum=checksum,
    )


def _required(value: SourcedValue | None, name: str) -> SourcedValue:
    if value is None:
        raise AircraftWorkbookError(f"required value {name!r} is blank")
    return value


def _performance_points(sheet: Any) -> tuple[AircraftPerformancePoint, ...]:
    expected = [
        "Weight (lb)",
        "Altitude (ft)",
        "Mach",
        "Cruise Allowed (Y/N)",
        "Fuel Burn (lb/hr)",
        "Available Thrust (lbf)",
        "Source",
        "Page/Figure",
        "Notes",
    ]
    headers = [cell.value for cell in sheet[1]][:9]
    if headers != expected:
        raise AircraftWorkbookError("Performance_Map columns do not match the published contract")
    points: list[AircraftPerformancePoint] = []
    coordinates: set[tuple[float, float, float]] = set()
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row[:9])
        if not any(value not in (None, "") for value in values):
            continue
        # The distributed template contains an instructional Tier row, not engineering data.
        if all(isinstance(value, str) for value in values[:4]):
            continue
        required = values[:4] + [values[6]]
        if any(value in (None, "") for value in required):
            raise AircraftWorkbookError(
                f"Performance_Map row {index} is partial; complete it or omit it"
            )
        try:
            weight_kg = float(to_si(float(values[0]), "lb")[0])
            altitude_m = float(to_si(float(values[1]), "ft")[0])
            mach = float(values[2])
        except (TypeError, ValueError) as exc:
            raise AircraftWorkbookError(
                f"Performance_Map row {index} has invalid numeric data"
            ) from exc
        coordinate = (weight_kg, altitude_m, mach)
        if coordinate in coordinates:
            raise AircraftWorkbookError(f"duplicate performance grid coordinate at row {index}")
        coordinates.add(coordinate)
        allowed_text = str(values[3]).strip().upper()
        if allowed_text not in {"Y", "N"}:
            raise AircraftWorkbookError(f"Performance_Map row {index} cruise flag must be Y or N")
        if not values[7]:
            warnings.warn(f"Performance_Map row {index} has no page/figure", stacklevel=2)
        fuel = None if values[4] in (None, "") else float(to_si(float(values[4]), "lb/hr")[0])
        thrust = None if values[5] in (None, "") else float(to_si(float(values[5]), "lbf")[0])
        points.append(
            AircraftPerformancePoint(
                weight_kg=weight_kg,
                altitude_m=altitude_m,
                mach=mach,
                cruise_allowed=allowed_text == "Y",
                fuel_burn_kg_s=fuel,
                available_thrust_n=thrust,
                source=str(values[6]),
                page_figure=str(values[7]) if values[7] else None,
                notes=str(values[8]) if values[8] else None,
            )
        )
    return tuple(points)


def load_aircraft_workbook(path: str | Path) -> AircraftModel:
    """Load and strictly validate an aircraft workbook while retaining traceability."""

    workbook_path = Path(path)
    if not workbook_path.exists():
        raise AircraftWorkbookError(
            f"aircraft workbook not found at {workbook_path}; copy it to "
            "aircraft_database/NASA_STCA_55T/NASA_STCA_55T_Aircraft.xlsx"
        )
    checksum = sha256_file(workbook_path)
    retrieved_at = datetime.now(UTC)
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    missing = REQUIRED_SHEETS - set(workbook.sheetnames)
    if missing:
        raise AircraftWorkbookError(f"missing required sheets: {', '.join(sorted(missing))}")
    general_rows = _parameter_rows(workbook["General"])
    limit_rows = _parameter_rows(workbook["Operating_Limits"])

    def general(name: str) -> SourcedValue | None:
        row = general_rows.get(name)
        return (
            None
            if row is None
            else _sourced(
                row,
                parameter=name,
                workbook=workbook_path,
                checksum=checksum,
                retrieved_at=retrieved_at,
            )
        )

    def limit(name: str) -> SourcedValue | None:
        row = limit_rows.get(name)
        return (
            None
            if row is None
            else _sourced(
                row,
                parameter=name,
                workbook=workbook_path,
                checksum=checksum,
                retrieved_at=retrieved_at,
            )
        )

    dimensions = {
        key: value
        for key in ("Length", "Wingspan", "Height", "Wing Area", "Number of Engines")
        if (value := general(key)) is not None
    }
    return AircraftModel(
        name=_required(general("Aircraft Name"), "Aircraft Name"),
        manufacturer=_required(general("Manufacturer"), "Manufacturer"),
        variant=general("Variant"),
        dimensions=dimensions,
        operating_limits=AircraftOperatingLimits(
            mtow=_required(limit("MTOW"), "MTOW"),
            oew=_required(limit("OEW"), "OEW"),
            maximum_operating_mach=_required(
                limit("Maximum Operating Mach"), "Maximum Operating Mach"
            ),
            maximum_cruise_mach=_required(limit("Maximum Cruise Mach"), "Maximum Cruise Mach"),
            minimum_sustained_supersonic_mach=_required(
                limit("Minimum Sustained Supersonic Mach"), "Minimum Sustained Supersonic Mach"
            ),
            service_ceiling=_required(limit("Service Ceiling"), "Service Ceiling"),
            minimum_cruise_altitude=limit("Minimum Cruise Altitude"),
        ),
        performance_map=_performance_points(workbook["Performance_Map"]),
        workbook_checksum=checksum,
    )
