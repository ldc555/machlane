"""Fail-closed readiness audit for the physical sonic-boom calculation path."""

from __future__ import annotations

from enum import StrEnum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pydantic import Field

from open_mco.aircraft import (
    AircraftWorkbookError,
    load_aircraft_definition_workbook,
    load_aircraft_workbook,
)
from open_mco.aircraft.loader import REQUIRED_SHEETS, sha256_file
from open_mco.models import FrozenModel
from open_mco.physics.signatures import NearFieldSignatureError, read_near_field_samples


class ReadinessStatus(StrEnum):
    READY = "READY"
    MISSING_INPUT = "MISSING_INPUT"
    INVALID = "INVALID"
    NOT_IMPLEMENTED = "NOT_IMPLEMENTED"


class ReadinessCheck(FrozenModel):
    key: str
    status: ReadinessStatus
    detail: str
    action: str | None = None


class BoomReadinessReport(FrozenModel):
    workbook_path: str
    workbook_checksum: str | None
    ready_for_physical_prediction: bool
    checks: tuple[ReadinessCheck, ...] = Field(min_length=1)

    @property
    def blockers(self) -> tuple[ReadinessCheck, ...]:
        return tuple(check for check in self.checks if check.status != ReadinessStatus.READY)


def _parameter_values(sheet: Any) -> dict[str, Any]:
    return {
        str(row[0]).strip(): row[1]
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row[0] not in (None, "")
    }


def _required_blanks(sheet: Any) -> list[str]:
    missing: list[str] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[3] or "").strip().lower() == "yes" and row[1] in (None, ""):
            missing.append(str(row[0]).strip())
    return missing


def _has_performance_data(sheet: Any) -> bool:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = list(row[:9])
        if not any(value not in (None, "") for value in values):
            continue
        if all(isinstance(value, str) for value in values[:4]):
            continue
        return True
    return False


def assess_boom_readiness(
    workbook_path: str | Path, *, near_field_path: str | Path | None = None
) -> BoomReadinessReport:
    """Inventory every current blocker without invoking weather, terrain, CFD, or PCBoom."""

    path = Path(workbook_path)
    if not path.exists():
        return BoomReadinessReport(
            workbook_path=str(path),
            workbook_checksum=None,
            ready_for_physical_prediction=False,
            checks=(
                ReadinessCheck(
                    key="aircraft_workbook",
                    status=ReadinessStatus.MISSING_INPUT,
                    detail=f"Workbook not found: {path}",
                    action="Provide the versioned aircraft workbook.",
                ),
            ),
        )

    checksum = sha256_file(path)
    workbook = load_workbook(path, data_only=True, read_only=True)
    current_definition = None
    if "Sonic_Boom" in workbook.sheetnames and "Nearfield_Signatures" in workbook.sheetnames:
        try:
            current_definition = load_aircraft_definition_workbook(path)
        except AircraftWorkbookError:
            current_definition = None
    missing_sheets = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
    if missing_sheets:
        return BoomReadinessReport(
            workbook_path=str(path.resolve()),
            workbook_checksum=checksum,
            ready_for_physical_prediction=False,
            checks=(
                ReadinessCheck(
                    key="aircraft_workbook_contract",
                    status=ReadinessStatus.INVALID,
                    detail=f"Missing sheets: {', '.join(missing_sheets)}",
                    action="Restore the published workbook structure.",
                ),
            ),
        )

    checks: list[ReadinessCheck] = [
        ReadinessCheck(
            key="aircraft_workbook_contract",
            status=ReadinessStatus.READY,
            detail="Required sheets are present and the workbook is checksummed.",
        )
    ]
    missing_aircraft = _required_blanks(workbook["General"]) + _required_blanks(
        workbook["Operating_Limits"]
    )
    if missing_aircraft:
        checks.append(
            ReadinessCheck(
                key="aircraft_required_data",
                status=ReadinessStatus.MISSING_INPUT,
                detail=f"Blank required fields: {', '.join(missing_aircraft)}",
                action="Populate reviewed values with source and page/figure traceability.",
            )
        )
    else:
        try:
            if current_definition is None:
                load_aircraft_workbook(path)
        except AircraftWorkbookError as exc:
            checks.append(
                ReadinessCheck(
                    key="aircraft_required_data",
                    status=ReadinessStatus.INVALID,
                    detail=str(exc),
                    action="Correct the workbook value or unit.",
                )
            )
        else:
            checks.append(
                ReadinessCheck(
                    key="aircraft_required_data",
                    status=ReadinessStatus.READY,
                    detail="Required aircraft values pass strict unit and provenance loading.",
                )
            )

    has_performance = _has_performance_data(workbook["Performance_Map"])
    checks.append(
        ReadinessCheck(
            key="aircraft_performance_map",
            status=ReadinessStatus.READY if has_performance else ReadinessStatus.MISSING_INPUT,
            detail=(
                "At least one performance operating point is populated."
                if has_performance
                else "No performance operating points are populated."
            ),
            action=None if has_performance else "Add reviewed weight/altitude/Mach performance rows.",
        )
    )

    mission_values = _parameter_values(workbook["Mission_Config"])
    mission_missing = [
        name
        for name in ("Required Reliability", "Boom Limit")
        if mission_values.get(name) in (None, "")
    ]
    declared_limit = mission_values.get("Boom Limit")
    limit_mismatch = False
    if not mission_missing:
        try:
            limit_mismatch = abs(float(str(declared_limit)) - 0.11) > 1e-9
        except (TypeError, ValueError):
            limit_mismatch = True
    checks.append(
        ReadinessCheck(
            key="mission_acceptance_inputs",
            status=(
                ReadinessStatus.MISSING_INPUT
                if mission_missing
                else ReadinessStatus.INVALID
                if limit_mismatch
                else ReadinessStatus.READY
            ),
            detail=(
                f"Blank mission settings: {', '.join(mission_missing)}"
                if mission_missing
                else (
                    f"Workbook boom limit is {declared_limit} psf; the current FAA NPRM "
                    "research threshold is 0.11 psf."
                    if limit_mismatch
                    else "Reliability target and 0.11 psf research threshold are populated."
                )
            ),
            action=(
                "Populate reviewed mission settings and units."
                if mission_missing
                else (
                    "Keep benchmark thresholds separate and set operational screening to 0.11 psf."
                    if limit_mismatch
                    else None
                )
            ),
        )
    )

    signature_path = Path(near_field_path) if near_field_path else None
    if signature_path is None and "Sonic_Boom_Optional" in workbook.sheetnames:
        boom_values = _parameter_values(workbook["Sonic_Boom_Optional"])
        declared = boom_values.get("Nearfield Signature File")
        if declared not in (None, ""):
            signature_path = Path(str(declared))
            if not signature_path.is_absolute():
                signature_path = path.parent / signature_path
    if (
        signature_path is None
        and current_definition is not None
        and current_definition.nearfield_samples
    ):
        operating_points = {
            (sample.mach, sample.altitude_ft, sample.reference_distance_ft)
            for sample in current_definition.nearfield_samples
        }
        checks.append(
            ReadinessCheck(
                key="near_field_signature",
                status=ReadinessStatus.READY,
                detail=(
                    f"{len(current_definition.nearfield_samples):,} embedded near-field samples "
                    f"cover {len(operating_points)} declared operating point(s)."
                ),
                action="Fail closed whenever the requested aircraft state is outside this coverage.",
            )
        )
        if len(operating_points) == 1:
            checks.append(
                ReadinessCheck(
                    key="near_field_operating_envelope",
                    status=ReadinessStatus.MISSING_INPUT,
                    detail=(
                        "The embedded near-field data covers only one Mach/altitude/reference-"
                        "distance operating point. It supports the LM1021 benchmark, not a "
                        "continuous climb/cruise/descent route envelope."
                    ),
                    action=(
                        "Add reviewed signatures for every supersonic operating state or constrain "
                        "physical analysis to the exact Mach 1.6 / 55,000 ft benchmark point."
                    ),
                )
            )
    elif signature_path is None:
        checks.append(
            ReadinessCheck(
                key="near_field_signature",
                status=ReadinessStatus.MISSING_INPUT,
                detail="No near-field pressure-signature file is declared.",
                action="Generate one with reviewed CFD (for example SU2) or provide measured data.",
            )
        )
    else:
        try:
            read_near_field_samples(signature_path)
        except NearFieldSignatureError as exc:
            checks.append(
                ReadinessCheck(
                    key="near_field_signature",
                    status=ReadinessStatus.INVALID,
                    detail=str(exc),
                    action="Export the signature using the documented SI CSV contract.",
                )
            )
        else:
            checks.append(
                ReadinessCheck(
                    key="near_field_signature",
                    status=ReadinessStatus.READY,
                    detail=f"Signature schema and samples are valid: {signature_path}",
                )
            )

    if current_definition is not None:
        benchmark_ids = {
            profile.profile_id for profile in current_definition.benchmark_atmospheres
        }
        benchmark_ready = "nasa_profile_1" in benchmark_ids
        checks.append(
            ReadinessCheck(
                key="propagation_benchmark_atmospheres",
                status=(
                    ReadinessStatus.READY
                    if benchmark_ready
                    else ReadinessStatus.MISSING_INPUT
                ),
                detail=(
                    f"Imported {len(benchmark_ids)} NASA propagation benchmark atmosphere(s)."
                    if benchmark_ready
                    else "NASA LM1021 required atmosphere profile 1 is not embedded."
                ),
                action=(
                    None
                    if benchmark_ready
                    else "Import NASA atmosphere profile 1 before validating the solver."
                ),
            )
        )

    checks.extend(
        (
            ReadinessCheck(
                key="atmospheric_column",
                status=ReadinessStatus.READY,
                detail="HRRR, GEFS, and ERA5 normalize temperature, pressure, wind, and humidity fields.",
            ),
            ReadinessCheck(
                key="terrain_profile",
                status=ReadinessStatus.READY,
                detail="USGS 3DEP and local-raster adapters normalize terrain profiles.",
            ),
            ReadinessCheck(
                key="physical_propagation_engine",
                status=ReadinessStatus.NOT_IMPLEMENTED,
                detail="No reviewed nonlinear ray/waveform propagation engine is registered.",
                action="Integrate and validate a physical engine through SonicBoomPropagationEngine.",
            ),
            ReadinessCheck(
                key="reference_validation",
                status=ReadinessStatus.MISSING_INPUT,
                detail="The PCBoom offline adapter exists, but no comparison result is supplied.",
                action="Run a declared comparison matrix under a separate NASA software agreement.",
            ),
        )
    )
    return BoomReadinessReport(
        workbook_path=str(path.resolve()),
        workbook_checksum=checksum,
        ready_for_physical_prediction=all(
            check.status == ReadinessStatus.READY for check in checks
        ),
        checks=tuple(checks),
    )
