from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

PARAMETER_HEADER = ["Parameter", "Value", "Unit", "Required", "Source", "Page/Figure", "Notes"]


@pytest.fixture
def aircraft_workbook(tmp_path: Path) -> Path:
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme.append(["Aircraft Data Specification v1.1"])

    general = workbook.create_sheet("General")
    general.append(PARAMETER_HEADER)
    general_rows = [
        ["Aircraft Name", "Test SST", "-", "Yes", "Test report", "p. 1", None],
        ["Manufacturer", "Example", "-", "Yes", "Test report", "p. 1", None],
        ["Variant", "A", "-", "No", "Test report", "p. 1", None],
        ["Length", 200, "ft", "Yes", "Test report", "p. 2", None],
        ["Wingspan", 80, "ft", "Yes", "Test report", "p. 2", None],
        ["Wing Area", 3000, "ft²", "Yes", "Test report", "p. 2", None],
        ["Number of Engines", 2, "count", "Yes", "Test report", "p. 3", None],
    ]
    for row in general_rows:
        general.append(row)

    limits = workbook.create_sheet("Operating_Limits")
    limits.append(PARAMETER_HEADER)
    limit_rows = [
        ["MTOW", 120000, "lb", "Yes", "Test report", "p. 4", None],
        ["OEW", 70000, "lb", "Yes", "Test report", "p. 4", None],
        ["Maximum Operating Mach", 1.8, "Mach", "Yes", "Test report", "p. 5", None],
        ["Maximum Cruise Mach", 1.6, "Mach", "Yes", "Test report", "p. 5", None],
        ["Minimum Sustained Supersonic Mach", 1.0, "Mach", "Yes", "Test report", "p. 5", None],
        ["Service Ceiling", 55000, "ft", "Yes", "Test report", "p. 5", None],
        ["Minimum Cruise Altitude", 35000, "ft", "No", "Test report", "p. 5", None],
    ]
    for row in limit_rows:
        limits.append(row)

    performance = workbook.create_sheet("Performance_Map")
    performance.append(
        [
            "Weight (lb)",
            "Altitude (ft)",
            "Mach",
            "Cruise Allowed (Y/N)",
            "Fuel Burn (lb/hr)",
            "Available Thrust (lbf)",
            "Source",
            "Page/Figure",
            "Notes",
        ]
    )
    performance.append([100000, 45000, 1.2, "Y", 8000, 20000, "Test report", "p. 6", "test"])
    workbook.create_sheet("Aerodynamics_Optional").append(
        ["Mach", "CL", "CD", "L/D", "Source", "Page/Figure", "Notes"]
    )
    sonic = workbook.create_sheet("Sonic_Boom_Optional")
    sonic.append(PARAMETER_HEADER)
    mission = workbook.create_sheet("Mission_Config")
    mission.append(PARAMETER_HEADER)
    mission.append(["Required Reliability", 0.95, "probability", "No", "Test", "p. 7", None])

    path = tmp_path / "aircraft.xlsx"
    workbook.save(path)
    return path
