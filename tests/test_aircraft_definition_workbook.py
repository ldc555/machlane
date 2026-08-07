from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from open_mco.aircraft import (
    AircraftWorkbookError,
    export_aircraft_definition_workbook,
    load_aircraft_definition_workbook,
)

PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = PROJECT_ROOT / "aircraft_database/Aircraft_One/Aircraft_One_Template.xlsx"


def _legacy_nasa_payload() -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    parameter_headers = [
        "Parameter",
        "Value",
        "Unit",
        "Required",
        "Source",
        "Page/Figure",
        "Notes",
    ]
    source = "NASA STCA report\nhttps://ntrs.nasa.gov/citations/20200000513"
    for name, rows in {
        "General": [
            ["Aircraft Name", "NASA 55t STCA", "-", "Yes", source, "Table 2", None],
            ["Manufacturer", "NASA", "-", "Yes", source, "p. 1", None],
        ],
        "Operating_Limits": [
            ["MTOW", 121_000, "lb", "Yes", source, "Table 2", None],
            ["Maximum Cruise Mach", 1.4, "Mach", "Yes", source, "Table 2", None],
        ],
        "Mission_Config": [
            ["Preferred Cruise Mach", 1.4, "Mach", "No", source, "Table 2", None],
            ["Preferred Cruise Altitude", 50_000, "ft", "No", source, "Table 3", None],
        ],
        "Sonic_Boom_Optional": [
            ["Nearfield Signature File", None, "file", "No", None, None, None],
        ],
    }.items():
        sheet = workbook.create_sheet(name)
        sheet.append(parameter_headers)
        for row in rows:
            sheet.append(row)
    performance = workbook.create_sheet("Performance_Map")
    performance.append(
        [
            "Weight (lb)",
            "Altitude (ft)",
            "Mach",
            "Cruise Allowed (Y/N)",
            "Fuel Burn (lb/hr)",
            "Available Thrust (lbf)",
            "Source",
            "Page/Figure",
            "Notes",
        ]
    )
    profile = workbook.create_sheet("Phase_Profile")
    profile.append(
        ["Sequence", "Phase", "Altitude (ft)", "Mach", "Source", "Page/Figure", "Notes"]
    )
    profile.append([1, "Takeoff", 0, 0.3, source, "Table 1", None])
    profile.append([2, "Supersonic cruise", 50_000, 1.4, source, "Table 1", None])
    profile.append([3, "Approach", 5_000, 0.4, source, "Table 1", None])
    timing = workbook.create_sheet("Phase_Timing")
    timing.append(["Phase", "Duration (min)", "Basis", "Source", "Page/Figure", "Notes"])
    for phase, duration, basis in (
        ("taxi_out", 9, "NASA_STCA"),
        ("climb_acceleration", 47, "NASA_STCA"),
        ("cruise", None, "CALCULATED"),
        ("descent", 20, "NASA_STCA"),
        ("approach", 4, "NASA_STCA"),
        ("taxi_in", 5, "NASA_STCA"),
    ):
        timing.append([phase, duration, basis, source if duration else None, "Table 3" if duration else None, None])
    payload = BytesIO()
    workbook.save(payload)
    return payload.getvalue()


def test_aircraft_one_template_imports_without_inventing_missing_models() -> None:
    aircraft = load_aircraft_definition_workbook(TEMPLATE)

    assert aircraft.aircraft_id == "aircraft_one"
    assert aircraft.display_name == "Aircraft One"
    assert aircraft.value("Aircraft Name") == "Boom XB-1"
    assert aircraft.numeric_value("Maximum Demonstrated Mach") == 1.18
    assert aircraft.value("Maximum Operating Mach") is None
    assert len(aircraft.phase_timing) == 6
    assert not aircraft.phase_profile_ready
    assert not aircraft.performance_data_ready
    assert not aircraft.nearfield_ready
    assert aircraft.workbook_checksum is not None


def test_legacy_nasa_workbook_auto_detects_and_populates() -> None:
    aircraft = load_aircraft_definition_workbook(_legacy_nasa_payload())

    assert aircraft.display_name == "NASA 55t STCA"
    assert aircraft.value("Aircraft Name") == "NASA 55t STCA"
    assert aircraft.numeric_value("MTOW") == 121_000
    assert aircraft.numeric_value("Preferred Cruise Mach") == 1.4
    assert aircraft.phase_profile_ready
    assert not aircraft.performance_data_ready
    assert not aircraft.nearfield_ready


def test_normalized_aircraft_excel_round_trips() -> None:
    original = load_aircraft_definition_workbook(_legacy_nasa_payload())

    payload = export_aircraft_definition_workbook(original)
    restored = load_aircraft_definition_workbook(payload)

    assert restored.aircraft_id == "aircraft_one"
    assert restored.value("Aircraft Name") == "NASA 55t STCA"
    assert restored.numeric_value("MTOW") == 121_000
    assert restored.phase_profile == original.phase_profile
    assert restored.phase_timing == original.phase_timing
    assert restored.performance_map == original.performance_map
    assert restored.nearfield_samples == original.nearfield_samples


def test_aircraft_one_import_rejects_a_missing_contract_sheet() -> None:
    workbook = load_workbook(TEMPLATE)
    del workbook["Nearfield_Signatures"]
    payload = BytesIO()
    workbook.save(payload)

    with pytest.raises(AircraftWorkbookError, match="Nearfield_Signatures"):
        load_aircraft_definition_workbook(payload.getvalue())


def test_aircraft_one_import_rejects_a_partial_performance_row() -> None:
    workbook = load_workbook(TEMPLATE)
    sheet = workbook["Performance_Map"]
    sheet["A2"] = 13_500
    sheet["B2"] = 35_000
    payload = BytesIO()
    workbook.save(payload)

    with pytest.raises(AircraftWorkbookError, match="partial"):
        load_aircraft_definition_workbook(payload.getvalue())
