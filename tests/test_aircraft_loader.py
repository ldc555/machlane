from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from open_mco.aircraft import AircraftWorkbookError, load_aircraft_workbook


def test_loads_and_normalizes_aircraft(aircraft_workbook: Path) -> None:
    aircraft = load_aircraft_workbook(aircraft_workbook)
    assert aircraft.name.original_value == "Test SST"
    assert aircraft.dimensions["Length"].value_si == pytest.approx(60.96)
    assert aircraft.operating_limits.mtow.si_unit == "kg"
    assert aircraft.performance_map[0].fuel_burn_kg_s == pytest.approx(1.00798, rel=1e-4)
    assert len(aircraft.workbook_checksum) == 64


def test_missing_workbook_has_copy_instruction(tmp_path: Path) -> None:
    with pytest.raises(AircraftWorkbookError, match="copy it to"):
        load_aircraft_workbook(tmp_path / "missing.xlsx")


def test_required_blank_and_unknown_units_fail(aircraft_workbook: Path) -> None:
    workbook = load_workbook(aircraft_workbook)
    workbook["General"]["B2"] = None
    workbook.save(aircraft_workbook)
    with pytest.raises(AircraftWorkbookError, match="Aircraft Name.*blank"):
        load_aircraft_workbook(aircraft_workbook)

    workbook = load_workbook(aircraft_workbook)
    workbook["General"]["B2"] = "Test SST"
    workbook["General"]["C5"] = "furlongs_per_fortnight"
    workbook.save(aircraft_workbook)
    with pytest.raises(AircraftWorkbookError, match="unknown or incompatible unit"):
        load_aircraft_workbook(aircraft_workbook)


def test_partial_and_duplicate_performance_rows_fail(aircraft_workbook: Path) -> None:
    workbook = load_workbook(aircraft_workbook)
    workbook["Performance_Map"].append([100000, 45000, 1.3, None, None, None, "Test report"])
    workbook.save(aircraft_workbook)
    with pytest.raises(AircraftWorkbookError, match="partial"):
        load_aircraft_workbook(aircraft_workbook)

    workbook = load_workbook(aircraft_workbook)
    workbook["Performance_Map"].delete_rows(3)
    workbook["Performance_Map"].append([100000, 45000, 1.2, "Y", 8000, 20000, "Test", "p.6"])
    workbook.save(aircraft_workbook)
    with pytest.raises(AircraftWorkbookError, match="duplicate"):
        load_aircraft_workbook(aircraft_workbook)


def test_populated_value_without_traceability_warns(aircraft_workbook: Path) -> None:
    workbook = load_workbook(aircraft_workbook)
    workbook["General"]["E2"] = None
    workbook.save(aircraft_workbook)
    with pytest.warns(UserWarning, match="traceability"):
        load_aircraft_workbook(aircraft_workbook)
